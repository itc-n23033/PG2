import openpyxl
import sys

def read_excel_file(file_path):
    """
    Excelファイルからデータを読み込みます。

    :param file_path: Excelファイルのパス
    :return: データのリスト
    """
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))
    return data

def transpose_data(data):
    """
    データを転置します。

    :param data: データのリスト
    :return: 転置されたデータのリスト
    """
    transposed_data = list(map(list, zip(*data)))
    return transposed_data

def write_excel_file(data, file_path):
    """
    データをExcelファイルに書き込みます。

    :param data: データのリスト
    :param file_path: Excelファイルのパス
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    for row in data:
        sheet.append(row)
    wb.save(file_path)

if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("使用法: python script.py 入力ファイルパス 出力ファイルパス")
        sys.exit(1)
    
    input_file_path = sys.argv[1]
    output_file_path = sys.argv[2]
    
    data = read_excel_file(input_file_path)
    transposed_data = transpose_data(data)
    write_excel_file(transposed_data, output_file_path)

    print('完了')
