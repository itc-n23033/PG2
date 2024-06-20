import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import sys

def create_multiplication_table(num):
    multiplication_table = [[i * j for j in range(1, num + 1)] for i in range(1, num + 1)]
    return multiplication_table

def write_multiplication_table_to_excel(multiplication_table, num):
    wb = Workbook()
    ws = wb.active
    ws.title = "Multiplication Table"

    bold_font = Font(bold=True)

    # 見出しを追加し、太字にする
    for col_num in range(1, num + 1):
        cell = ws.cell(row=1, column=col_num + 1, value=col_num)
        cell.font = bold_font
        cell = ws.cell(row=col_num + 1, column=1, value=col_num)
        cell.font = bold_font

    # 値を追加
    for i, row in enumerate(multiplication_table):
        for j, value in enumerate(row):
            ws.cell(row=i + 2, column=j + 2, value=value)

    wb.save("multiplication_table.xlsx")

# コマンドライン引数から入力を受け取る
if len(sys.argv) != 2:
    print("使用法: python multiplication_table.py <N>")
    sys.exit(1)

try:
    num = int(sys.argv[1])
    if num < 1:
        raise ValueError("1以上の整数を入力してください。")
except ValueError as e:
    print(e)
    sys.exit(1)

multiplication_table = create_multiplication_table(num)
write_multiplication_table_to_excel(multiplication_table, num)

print('完了')

